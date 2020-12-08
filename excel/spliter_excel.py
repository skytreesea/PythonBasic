#문장 나눠서 엑셀파일로 송출하기, 기본형 폴더명만 바꿔서 사용하면 됨, re.split 실습, openpyxl 실습
import os
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Font, Border, Side
import re, pickle
# d로 먼저 문장쪼개기 연습
d='In addition, exercise can help you sleep better at night., In addition, when you laugh, your brain works better., We built a room addition to our house.,'
# b는 한국어 c는 영어: 문장을 쪼개서 엑셀로 저장
b='그는 아침을 먹는다 .,그는 그의 집을 떠난다. ,그는 버스 정류장으로 걸어 간다. ,그는 5 분을 기다린다. ,그는 버스에 탄다. ,그는 4 분의 1에 넣는다. ,그는 앉았다. ,그는 창 밖을 본다. ,그는 그의 학교를보고있다. ,그는 버스 문으로 걸어 간다. ,그는 버스 운전사에게 감사드립니다. ,그는 버스를 빠져 나간다.'
c='He leaves his house., He walks to the bus stop., He eats his breakfast., He waits five minutes., He gets on the bus., He puts in a quarter., He sits down., He looks outside the window., He sees his school., He walks to the bus door., He thanks the bus driver., He exits the bus.'
print(re.split(r'[,]', d))

def sent_tokenize(text):
    sentences = re.split(r"[.!?]", text)
    sentences = [sent.strip(" ") for sent in sentences]
    return sentences

sb=sent_tokenize(b)
sc=sent_tokenize(c)

wb = Workbook()
ws = wb.active
k=0
while k < len(sb):
	ws['A'+str(k+1)] = sb[k]
	ws['B'+str(k+1)] = sc[k]
	ws.column_dimensions['A'].width= 40
	ws.column_dimensions['B'].width= 36		
	k +=1
#폴더 명만 지정
os.chdir(r'C:\Users\ERC\Documents\GitHub\PythonBasic\excel')
#엑셀 저장
wb.save('spliter_excel_20201209.xlsx')
